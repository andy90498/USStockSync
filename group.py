import openpyxl
import os
from config import EXCEL_PATH, CREDENTIALS_PATH, BASE, APPDATA
from openpyxl.utils import get_column_letter
#宣告一下，怕等一下忘記
SHEET_NAME = "Group"

def ensure_group_sheet():
    #如果EXCEL檔案不存在
    if not os.path.exists(EXCEL_PATH):
        wb = openpyxl.Workbook()
        wb.save(EXCEL_PATH)
    wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)
    #如果分頁不存在
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        ws.cell(row=1, column=1, value="序號")  # A欄作為標示保留
        ws.cell(row=1, column=2, value="Default Group")  # B欄為預設群組
    wb.save(EXCEL_PATH)

def save_group_data(groups_dict):
    #將整個群組資料結構同步到Excel

    ensure_group_sheet()
    wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)
    ws = wb[SHEET_NAME]

    ws.delete_cols(2, ws.max_column)  # 清空舊資料（保留 A 欄）

    col = 2
    #從B欄開始迴圈到groups_dict結束，將字典中每個項目存到row1作為群組名稱
    for group_name, stock_list in groups_dict.items():
        ws.cell(row=1, column=col, value=group_name)
        #從row2開始，迴圈stock_list，直到字典結束，此處為股票代碼
        for idx, code in enumerate(stock_list, start=2):
            ws.cell(row=idx, column=col, value=code)
        col += 1
    wb.save(EXCEL_PATH)

def rename_group(old_name, new_name):
    ensure_group_sheet()
    wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)
    ws = wb[SHEET_NAME]

    for cell in ws[1]:
        if cell.value == old_name:
            cell.value = new_name
            break
    wb.save(EXCEL_PATH)

def delete_group(group_name):
    ensure_group_sheet()
    wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)
    ws = wb[SHEET_NAME]

    for col in range(2, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == group_name:
            ws.delete_cols(col, 1)
            break
    wb.save(EXCEL_PATH)



def update_single_group_column(group_name, stock_list):
    #徹底清空EXCEL中指定群組欄的資料並重寫
    ensure_group_sheet()
    wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)
    ws = wb[SHEET_NAME]

    target_col = None
    for col in range(2, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == group_name:
            target_col = col
            break

    if target_col is None:
        return

    #如果該欄有值，透過""的方式將每個儲存格清空
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=target_col, value="")

    #再寫入新的資料
    for idx, code in enumerate(stock_list, start=2):
        ws.cell(row=idx, column=target_col, value=code)

    wb.save(EXCEL_PATH)

def load_group_data():
    #從EXCEL載入所有群組與股票代碼
    ensure_group_sheet()
    wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)
    ws = wb[SHEET_NAME]

    groups = {}
    for col in range(2, ws.max_column + 1):
        group_name = ws.cell(row=1, column=col).value
        if not group_name:
            continue
        stock_list = []
        for row in range(2, ws.max_row + 1):
            code = ws.cell(row=row, column=col).value
            if code:
                stock_list.append(str(code).strip().upper())
        groups[group_name] = stock_list

    return groups