import requests
from bs4 import BeautifulSoup
import openpyxl
import psutil
import os
import time
import random
import datetime
import ctypes          # 用於關閉視窗
import subprocess
import tkinter.messagebox as messagebox
from config import EXCEL_PATH, CREDENTIALS_PATH, BASE, APPDATA

#建立一個陣列，簡列一些使用者瀏覽資訊，方便未來可以random讀取
USER_AGENTS = [
    # Windows 10 – Chrome
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko)"
    " Chrome/115.0.0.0 Safari/537.36",
    # Windows 11 – Firefox
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:116.0) Gecko/20100101 Firefox/116.0",
    # macOS – Safari
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_5) AppleWebKit/605.1.15"
    " (KHTML, like Gecko) Version/16.5 Safari/605.1.15",
    # Ubuntu – Chrome
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko)"
    " Chrome/115.0.0.0 Safari/537.36",
    # iPhone – Safari
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_4 like Mac OS X) AppleWebKit/605.1.15"
    " (KHTML, like Gecko) Version/17.4 Mobile/15E148 Safari/604.1",
    # Android – Chrome
    "Mozilla/5.0 (Linux; Android 14; Pixel 8) AppleWebKit/537.36 (KHTML, like Gecko)"
    " Chrome/115.0.0.0 Mobile Safari/537.36",
]
#語系，同上
LANGUAGES = [
    "en-US,en;q=0.9",
    "zh-TW,zh;q=0.8,en-US;q=0.5,en;q=0.3",
    "ja-JP,ja;q=0.9",
    "ko-KR,ko;q=0.9",
]

#嘗試存取指定檔案，open(,"a")可以在沒有檔案情況下自動建立檔案
def is_file_locked(EXCEL_PATH):
    if not os.path.exists(EXCEL_PATH):
        return False
    try:
        with open(EXCEL_PATH, "a"):
            return False
    except PermissionError:
        return True

#強制關閉執行程序(taskkill)，psutil.process_iter為目前執行中的程序陣列，透過迴圈for尋找name為excel.exe的程序並嘗試taskkill
def close_excel_instances(filename):
    closed = False
    for proc in psutil.process_iter(attrs=["pid", "name", "cmdline"]):
        try:
            #透過if單獨確認name那個欄位的值
            if "EXCEL.EXE" in proc.info["name"].upper():
                if any(filename.lower() in str(arg).lower() for arg in proc.info["cmdline"]):
                    proc.kill()
                    closed = True
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            pass
    return closed

#爬蟲母體之一，於Overview頁面(即該股票首頁)抓取Shares Out, PE Ratio, Price Target
#一開始先將這三個值設定為 - ，待後續抓到值後覆蓋 - 為正確資料

def fetch_overview_metrics(code, max_retries=5):
    url = f"https://stockanalysis.com/stocks/{code}/"
    shares_out = pe_ratio = price_target = "-"
    try:
        #發送request，並於失敗後自動重新發送，直到次數達到最大設定次數
        resp = safe_request(url, max_retries=max_retries)
        soup = BeautifulSoup(resp.text, "html.parser")
        #目標資料被藏在<td>底下
        tds = soup.find_all("td")
        #迴圈每個<td>，透過len取得最大td數量，因為我們要抓得值在被find的值+1的td中，所以在for迴圈時只能迴圈到len(tds)-1，避免溢位
        #而range則是從i = 0 to i = len(tds)-1
        #透過strip將頭尾空格去掉，再透過if比較關鍵字，符合條件後才存入對應變數
        for i in range(len(tds) - 1):
            lbl = tds[i].get_text(strip=True)
            val = tds[i+1].get_text(strip=True)
            if lbl.startswith("Shares Out"):
                shares_out = val
            elif lbl.startswith("PE Ratio"):
                pe_ratio = val
            elif lbl.startswith("Price Target"):
                price_target = val
    except Exception as e:
        #回傳錯誤代碼
        print(f"❌ fetch_overview_metrics({code}) 錯誤：{e}")
    return code, shares_out, pe_ratio, price_target



#爬蟲母體之一，於financials抓取EPS (Basic)、Free Cash Flow、Total Debt三種類別的資料
#因為這邊同一項目因年份關係，有六欄值，加上還要確認抓取到的年份，改用table而非td標籤
def fetch_financial_metrics(code, max_retries=5):
    url = f"https://stockanalysis.com/stocks/{code}/financials/"
    data = {}
    try:
        resp = safe_request(url, max_retries=max_retries)
        soup = BeautifulSoup(resp.text, "html.parser")
        table = soup.find("table")
        if not table:
            return data
        #年份在th標籤裡面
        years = [th.get_text(strip=True) for th in table.thead.find_all("th")][1:]
        #各項目在tbody的tr標籤中的td標籤裡面，透過迴圈搜尋tr標籤中的td標籤，並存入cols陣列中，
        for tr in table.tbody.find_all("tr"):
            cols = [td.get_text(strip=True) for td in tr.find_all("td")]
            lbl = cols[0] if cols else ""
            if lbl not in ["EPS (Basic)", "Free Cash Flow", "Total Debt"]:
                continue
                #透過zip將年分與抓到的數值綁定(cols[0]為標籤名稱，非數值)，並個別存到yr、val的變數中，遇到val的值為Upgrade用continue略過
            for yr, val in zip(years, cols[1:]):
                if val == "Upgrade":
                    continue
                #接著再接抓到的val數值對應進data字典中的同年份、同分類標籤底下，之後可以考慮data[{lbl}][{yr}]可能更好用
                data[f"{lbl} ({yr})"] = val
    except Exception as e:
        print(f"❌ fetch_financial_metrics({code}) 錯誤：{e}")
    return data




def fetch_total_debt(code, max_retries=5):
    """
    專門抓 Balance Sheet 裡的 Total Debt
    回傳 {"Total Debt (TTM)": "...", ...}
    """
    url = f"https://stockanalysis.com/stocks/{code}/financials/balance-sheet/"
    data = {}
    try:
        resp = safe_request(url, max_retries=max_retries)
        soup = BeautifulSoup(resp.text, "html.parser")
        #要抓六個年份，不適用單抓td
        table = soup.find("table")
        if not table:
            return data
        #從table中的th標籤抓年分 從[1]開始抓是因為[0]為其他字串，非年分
        years = [th.get_text(strip=True) for th in table.thead.find_all("th")][1:]
        #在tbody中迴圈每個tr標籤，並抓tr標籤中的td標籤存入cols陣列中
        for tr in table.tbody.find_all("tr"):
            cols = [td.get_text(strip=True) for td in tr.find_all("td")]
            #從cols陣列再去抓出標的關鍵字Total Debt
            if cols and cols[0] == "Total Debt":
                #綁定標的字串與數值並排除Upgrade地當次迴圈
                for yr, val in zip(years, cols[1:]):
                    if val == "Upgrade":
                        continue
                    #if true時代表資料正確，綁入data陣列中的{yr}變數下
                    data[f"Total Debt ({yr})"] = val
                break
    except Exception as e:
        print(f"❌ fetch_total_debt({code}) 錯誤：{e}")
    return data

#先找到excel中名為"股票代碼"的column，並繼續在該column中搜尋最靠近上方空白列的row座標
def get_next_available_row(ws, id_col_name="股票代碼", start_row=2):
    #找出股票代碼欄的column
    id_col = None
    for cell in ws[1]:
        if isinstance(cell.value, str) and cell.value.strip() == id_col_name:
            id_col = cell.column
            break
    if not id_col:
        id_col = 1  # fallback

    # 從指定列起往下找第一個股票代碼是空的row，
    for row in range(start_row, ws.max_row + 5):
        val = ws.cell(row=row, column=id_col).value
        if not str(val or "").strip():
            return row

    return ws.max_row + 1


def write_to_excel(data_rows, extra_data_dict=None, file_path=EXCEL_PATH, sheet_name="Crawl_data"):
    #檢查檔案是否開啟，透過def is_file_locked回傳布林
    auto_closed_excel = False
    if is_file_locked(EXCEL_PATH):
        user_choice = messagebox.askyesno(
            "Excel 檔案已開啟",
            f"{os.path.basename(EXCEL_PATH)} 檔案目前正在被 Excel 使用中。\n\n是否要強制關閉它來更新資料？（請確保檔案已儲存變更）"
        )
        if not user_choice:
            messagebox.showinfo("程序中止", "請先關閉Excel檔案後再執行。")
            return
        else:
            closed = close_excel_instances(os.path.basename(EXCEL_PATH))
            if closed:
                auto_closed_excel = True
            else:
                messagebox.showwarning("無法關閉", "系統無法自動關閉Excel，請手動關閉。")
                return

    #確保data_rows轉為陣列
    if isinstance(data_rows, tuple):
        data_rows = [data_rows]

    wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    header = {}
    for cell in ws[1]:
        val = cell.value
        if isinstance(val, str):
            header[val] = cell.column
        elif cell.data_type == "f":  # 如果是公式
            val = str(cell.value).strip()
            if cell.internal_value:
                header[cell.internal_value.strip()] = cell.column

    #先寫入第一頁爬到的資料
    for code, shares, pe, price_target in data_rows:
        if not code:
            continue

        # 找資料列
        found_row = None
        #如果strip的code與儲存格中相同，將該Row設為寫入目標
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=header.get("股票代碼", 1)).value).strip().upper() == code.strip().upper():
                found_row = row
                break
        #如果found_row有值則寫入row，如果沒有就取最靠近上方的空白列
        row = found_row if found_row else get_next_available_row(ws)

        # 主欄位寫入
        ws.cell(row=row, column=header.get("股票代碼", 1), value=code)
        ws.cell(row=row, column=header.get("SharesOut", 3), value=shares)
        ws.cell(row=row, column=header.get("PE Ratio", 4), value=pe)
        ws.cell(row=row, column=header.get("Price Target", 5), value=price_target)

        #第二頁爬到的資料（EPS & FCF）
        if extra_data_dict:
            raw_data = extra_data_dict.get(code, {})
            for col_name in header:
                #遇到下列關鍵字則略過
                if col_name in ["股票代碼", "SharesOut", "PE Ratio", "Price Target", "更新時間"]:
                    continue

                #EPS部分的處理 先找要搜尋網頁的關鍵字
                if col_name.startswith("EPS"):
                    #in(含)
                    if "TTM" in col_name:
                        lookup_key = "EPS (Basic) (TTM)"
                    else:
                        year = col_name[col_name.find("(")+1:col_name.find(")")]
                        lookup_key = f"EPS (Basic) (FY {year})"

                # Free Cash Flow
                elif col_name.startswith("Free Cash Flow"):
                    if "TTM" in col_name:
                        lookup_key = "Free Cash Flow (TTM)"
                    else:
                        year = col_name[col_name.find("(")+1:col_name.find(")")]
                        lookup_key = f"Free Cash Flow (FY {year})"

                # Total Debt
                elif col_name.startswith("Total Debt"):
                    if "TTM" in col_name:
                        lookup_key = "Total Debt (TTM)"
                    else:
                        year = col_name[col_name.find("(")+1:col_name.find(")")]
                        lookup_key = f"Total Debt (FY {year})"

                else:
                    continue

                val = raw_data.get(lookup_key, "-")
                #這行先寫入公司名稱(從main回傳過來的)
                ws.cell(row=row, column=header.get("公司名稱", 2), value=raw_data.get("公司名稱", "-"))
                #這行則是寫入各類別資料到本地EXCEL中
                ws.cell(row=row, column=header[col_name], value=val)

        # 寫入資料更新時間
        timestamp_col = header.get("更新時間", 11)  # 若找不到就 fallback 用第11欄
        ws.cell(row=row, column=timestamp_col, value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    wb.save(EXCEL_PATH)

    if auto_closed_excel:
        try:
            subprocess.Popen(["start", "", EXCEL_PATH], shell=True)
        except Exception as e:
            print(f"無法自動打開 Excel：{e}")

#因為出現429代碼，所以先弄個def確認代碼429後執行重新連接，backoff單位是秒
def safe_request(url, headers=None, max_retries=5, backoff=(3, 6)):
    for attempt in range(1, max_retries + 1):
        #呼叫下方的隨機headers取得不相同的瀏覽組合
        h = headers or get_random_headers()
        resp = requests.get(url, headers=h, timeout=10)
        if resp.status_code == 429:
            wait = random.uniform(*backoff)
            print(f"⚠️ [{attempt}/{max_retries}] 429 Too Many Requests for {url}，等待 {wait:.1f}秒")
            time.sleep(wait)
            continue
        resp.raise_for_status()
        return resp
    #最後一次仍錯誤，就直接 raise
    resp.raise_for_status()


#這邊的陣列寫在最上方
def get_random_headers():
    return {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept-Language": random.choice(LANGUAGES),
    }


